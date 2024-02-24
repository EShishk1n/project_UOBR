from datetime import datetime

from openpyxl.reader.excel import load_workbook
from os.path import getmtime
import datetime


def open_work_sheet():
    """Открывает файл. Нужна для искючения ошибки при запуске сайта при отсутствии файла в папке"""

    wb = load_workbook(filename='dvizhenie/uploads/Движение_БУ.xlsx')

    return wb['Движение_БУ']


def take_rigs_position_data(table_start_row: int, table_end_row: int) -> list:
    """Имортирует данные по окончанию бурения для каждой БУ из файла 'Движение_БУ'"""

    sheet = open_work_sheet()

    min_row = table_start_row
    max_row = table_end_row
    min_col = 4
    max_col = 6

    rigs_position_data = []

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        rig_position_data = {'number': row[0],
                             'field': row[1],
                             'end_date': row[2].date()}
        rigs_position_data.append(rig_position_data)

    return rigs_position_data


def take_pads_data(table_start_row: int, table_end_row: int) -> list:
    """Имортирует данные по КП из файла 'Движение_БУ'"""

    sheet = open_work_sheet()

    min_row = table_start_row
    max_row = table_end_row
    min_col = 9
    max_col = 19

    pads_data = []

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        if row[0] is not None:
            if row[8] is None:
                marker = 'нет'
            else:
                marker = row[8]
            pad_data = {'number': row[0],
                        'field': row[1],
                        'first_stage_date': row[2].date(),
                        'second_stage_date': row[3].date(),
                        'required_capacity': row[5],
                        'required_mud': row[6],
                        'marker': marker,
                        'gs_quantity': row[10],
                        'nns_quantity': row[9] - row[10]
                        }
            pads_data.append(pad_data)

    return pads_data


def take_file_cration_data() -> str | None:
    """Получает дату создания файла, если файла нет, возвращает None"""

    try:
        return datetime.datetime.fromtimestamp(getmtime('dvizhenie/uploads/Движение_БУ.xlsx')).strftime(
            '%d-%m-%Y, %H:%M')
    except FileNotFoundError:
        return None
