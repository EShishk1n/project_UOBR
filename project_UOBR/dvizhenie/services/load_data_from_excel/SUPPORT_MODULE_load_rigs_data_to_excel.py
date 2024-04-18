import datetime

from openpyxl.reader.excel import load_workbook

from dvizhenie.models import DrillingRig, type_of_DR, Contractor, Pad, RigPosition


def open_work_sheet_Razmery():
    """Открывает файл. Нужна для искючения ошибки при запуске сайта при отсутствии файла в папке"""

    wb = load_workbook(filename='dvizhenie/uploads/Размеры_БУ.xlsx')

    return wb['Лист1']


def take_rigs_data(table_start_row: str, table_end_row: str) -> list:
    """Имортирует данные по окончанию бурения для каждой БУ из файла 'Движение_БУ'"""

    sheet = open_work_sheet_Razmery()

    min_row = table_start_row
    max_row = table_end_row
    min_col = 1
    max_col = 4

    rigs_data = []

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        rig_data = {'type': row[2],
                    'number': row[1],
                    'contractor': row[0],
                    'mud': row[3], }
        rigs_data.append(rig_data)

    return rigs_data


def put_rigs_data(table_start_row: str, table_end_row: str):
    """Вставляет данные, полученные из таблицы в модель RigPosition"""

    rigs_data = take_rigs_data(table_start_row=table_start_row, table_end_row=table_end_row)

    for rig_data in rigs_data:
        type_of_DR_ = type_of_DR.objects.filter(type=rig_data['type'])[0]
        contractor = Contractor.objects.filter(contractor=rig_data['contractor'])[0]

        print(type_of_DR_, rig_data['number'], contractor, rig_data['mud'])

        DrillingRig(type=type_of_DR_, number=rig_data['number'], contractor=contractor,
                    mud=rig_data['mud']).save()


def take_rigs_positions_data(table_start_row: str, table_end_row: str) -> list:
    """Имортирует данные по окончанию бурения для каждой БУ из файла 'Движение_БУ'"""

    sheet = open_work_sheet_Razmery()

    min_row = table_start_row
    max_row = table_end_row
    min_col = 2
    max_col = 6

    rigs_data = []

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        rig_data = {'number': row[0],
                    'kust': row[4]}
        rigs_data.append(rig_data)

    return rigs_data


def put_rigs_positions_data(table_start_row: str, table_end_row: str):
    """Вставляет данные, полученные из таблицы в модель RigPosition"""

    rigs_data = take_rigs_positions_data(table_start_row=table_start_row, table_end_row=table_end_row)

    for rig_data in rigs_data:
        Rig = DrillingRig.objects.filter(number=rig_data['number'])[0]
        Kust = Pad.objects.filter(number=rig_data['kust'])[0]

        print(Rig, Kust)

        RigPosition(drilling_rig=Rig, pad=Kust, start_date=datetime.datetime.strptime('2022-01-01', '%Y-%m-%d'),
                    end_date=datetime.datetime.strptime('2024-12-30', '%Y-%m-%d')).save()
